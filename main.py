# main.py (VERSÃO OTIMIZADA COM LAYOUT RESPONSIVO E NOVAS FUNÇÕES DE CUSTO)

from openpyxl.styles import PatternFill, Font
from openpyxl import load_workbook
import sys
import os
import json
import pandas as pd
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QPushButton, QLabel, QTextEdit, 
                             QFileDialog, QProgressBar, QMessageBox, QGroupBox,
                             QFormLayout, QLineEdit, QComboBox, QTableWidget, 
                             QTableWidgetItem, QDialog, QInputDialog, QHeaderView,
                             QSplitter, QDialogButtonBox) # <<< IMPORTAÇÃO NECESSÁRIA >>>
from PyQt5.QtCore import Qt
try:
    import win32com.client
    import pythoncom
    PYWIN32_DISPONIVEL = True
except ImportError:
    PYWIN32_DISPONIVEL = False
    print("AVISO: 'pywin32' não encontrado. Geração de PDF do orçamento será pulada.")
    print("Para instalar, rode: pip install pywin32")
# <<< IMPORTAÇÕES DAS CLASSES ENCAPSULADAS >>>
from code_manager import CodeGenerator
from history_manager import HistoryManager
from history_dialog import HistoryDialog
from processing import ProcessThread
from nesting_dialog import NestingDialog
from dxf_engine import get_dxf_bounding_box # <<< IMPORTAÇÃO NECESSÁRIA >>>
from calculo_cortes import orquestrar_planos_de_corte

# =============================================================================
# ESTILO VISUAL DA APLICAÇÃO (QSS - Qt StyleSheet)
# =============================================================================
INOVA_PROCESS_STYLE = """
/* ================================================================================
    Estilo INOVA PROCESS (v8 - Tema Preto & Branco de Alto Contraste)
================================================================================ */

/* Passo 1: Fundo principal e cor de texto padrão */
QWidget {
    background-color: #111111; /* Preto suave para o fundo principal */
    color: #FFFFFF;          /* Branco puro para o texto */
    font-family: 'Segoe UI', Arial, sans-serif;
    font-size: 7pt; 
    border: none;
}

QLabel {
    color: #FFFFFF;
    background: transparent;
}

/* Divisores (Splitter) */
QSplitter::handle { background-color: #333333; }
QSplitter::handle:hover { background-color: #FFFFFF; }
QSplitter::handle:pressed { background-color: #CCCCCC; }

/* Passo 2: Contêineres com fundo de cinza escuro para contraste */
QGroupBox, QTableWidget, QListView {
    background-color: #222222; 
    border: 1px solid #444444; /* Borda cinza */
    border-radius: 8px;
}
QGroupBox {
    margin-top: 1em; 
    font-weight: bold;
}

/* Passo 3: Cor de Destaque - Títulos com fundo branco e texto preto */
QGroupBox::title {
    subcontrol-origin: margin;
    subcontrol-position: top center;
    padding: 2px 8px;
    background-color: #FFFFFF; /* Destaque invertido (branco) */
    color: #000000;          /* Destaque invertido (preto) */
    border-radius: 4px;
    font-weight: bold;
}

/* Passo 4: Campos de Input e ComboBox */
QLineEdit, QTextEdit, QComboBox, QDoubleSpinBox, QSpinBox {
    background-color: #111111;   /* Mesmo fundo do QWidget */
    border: 1px solid #444444;   /* Borda cinza */
    border-radius: 4px;
    padding: 4px; 
    color: #FFFFFF;
}
QLineEdit:focus, QTextEdit:focus, QComboBox:focus, QDoubleSpinBox:focus, QSpinBox:focus {
    border: 1px solid #FFFFFF; /* Foco com borda branca */
}
QLineEdit::placeholder {
    color: #777777;
}

/* Detalhes do ComboBox */
QComboBox::drop-down { border: none; }
QComboBox::down-arrow {
    image: url(C:/Users/mathe/Desktop/INOVA_PROCESS/down_arrow.png); /* Manter o ícone de seta (assumindo que seja branco) */
    width: 10px; height: 10px; margin-right: 8px;
}
QComboBox QAbstractItemView {
    background-color: #222222;
    border: 1px solid #FFFFFF;
    selection-background-color: #FFFFFF;
    selection-color: #000000;
    outline: 0px;
}

/* Passo 5: Botões Padrão com tons de cinza */
QPushButton {
    background-color: #333333;   /* Cinza médio */
    color: #DDDDDD;          /* Branco suave */
    font-weight: bold;
    padding: 4px 8px; 
    border-radius: 4px;
}
QPushButton:hover { background-color: #555555; } /* Cinza mais claro no hover */
QPushButton:pressed { background-color: #222222; }

/* Botão Primário (Invertido, como os títulos) */
QPushButton#primaryButton { background-color: #FFFFFF; color: #000000; }
QPushButton#primaryButton:hover { background-color: #CCCCCC; }

/* Botões de estado (Mantidos por usabilidade) */
QPushButton#successButton { background-color: #107C10; color: #FFFFFF; }
QPushButton#successButton:hover { background-color: #159d15; }
QPushButton#warningButton { background-color: #DCA307; color: #1A202C; }
QPushButton#warningButton:hover { background-color: #f0b92a; }

/* Passo 6: Tabela */
QTableWidget {
    gridline-color: #444444;
}
QHeaderView::section {
    background-color: #222222;
    color: #DDDDDD; /* Branco suave para não competir com os dados */
    padding: 4px;
    border: 1px solid #444444;
    font-weight: bold;
}
QTableWidget::item {
    color: #FFFFFF;
    font-size: 8pt;
    padding: 4px;
}
/* Seleção da tabela também usa o destaque invertido */
QTableWidget::item:selected {
    background-color: #FFFFFF;
    color: #000000;
}

/* Barra de Log */
QTextEdit#logExecution {
    font-family: 'Courier New', Courier, monospace;
    background-color: #222222;
    color: #4FD1C5; /* Mantive o ciano aqui para um look "terminal" clássico */
}

/* Barras de Rolagem */
QScrollBar:vertical { border: none; background: #222222; width: 12px; margin: 0; }
QScrollBar::handle:vertical { background: #444444; min-height: 20px; border-radius: 6px; }
QScrollBar::handle:vertical:hover { background: #666666; }
QScrollBar:horizontal { border: none; background: #222222; height: 12px; margin: 0; }
QScrollBar::handle:horizontal { background: #444444; min-width: 20px; border-radius: 6px; }
QScrollBar::handle:horizontal:hover { background: #666666; }
QScrollBar::add-line, QScrollBar::sub-line { border: none; background: none; }
"""

# =============================================================================
# CLASSE PRINCIPAL DA INTERFACE GRÁFICA
# =============================================================================
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Gerador de Desenhos Técnicos e DXF INP - NOROACO")
        self.setGeometry(100, 100, 1280, 850) 
        self.setMinimumSize(1100, 800)

        self.code_generator = CodeGenerator()
        self.history_manager = HistoryManager()
        
        self.colunas_df = ['nome_arquivo', 'forma', 'espessura', 'qtd', 'largura', 'altura', 'diametro', 'rt_base', 'rt_height', 'trapezoid_large_base', 'trapezoid_small_base', 'trapezoid_height', 'furos']
        self.colunas_df = ['nome_arquivo', 'forma', 'espessura', 'qtd', 'largura', 'altura', 'diametro', 'rt_base', 'rt_height', 'trapezoid_large_base', 'trapezoid_small_base', 'trapezoid_height', 'furos', 'dxf_path']
        self.manual_df = pd.DataFrame(columns=self.colunas_df)
        self.excel_df = pd.DataFrame(columns=self.colunas_df)
        self.furos_atuais = []
        self.project_directory = None

        self.initUI() # Chama o método que constrói a UI
        self.connect_signals() # Chama o método que conecta os eventos
        
        self.set_initial_button_state()
        self.update_dimension_fields(self.forma_combo.currentText())

    def initUI(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        # --- Layout Superior (Inputs e Furos) ---
        top_h_layout = QHBoxLayout()
        
        # <<< MUDANÇA ESTRUTURAL 1: PAINEL ESQUERDO COM LARGURA MÍNIMA >>>
        left_panel_widget = QWidget()
        left_v_layout = QVBoxLayout(left_panel_widget)
        left_v_layout.setContentsMargins(0,0,0,0) # Remove margens internas
        left_panel_widget.setMinimumWidth(450) # Impede o "esmagamento"

        # --- Grupo 1: Projeto ---
        project_group = QGroupBox("1. Projeto")
        project_layout = QVBoxLayout()
        self.start_project_btn = QPushButton("Iniciar Novo Projeto...")
        self.history_btn = QPushButton("Ver Histórico de Projetos")
        project_layout.addWidget(self.start_project_btn)
        project_layout.addWidget(self.history_btn)
        project_group.setLayout(project_layout)
        left_v_layout.addWidget(project_group)
        
        # --- (NOVO) Grupo 2: Parâmetros de Custo ---
        cost_group = QGroupBox("2. Parâmetros de Custo")
        cost_layout = QFormLayout()
        cost_layout.setLabelAlignment(Qt.AlignRight)
        # Valores padrão baseados na sua imagem
        self.imposto_input = QLineEdit("0,12") 
        self.frete_input = QLineEdit("0,26")
        cost_layout.addRow("Imposto (%):", self.imposto_input)
        cost_layout.addRow("Frete (R$):", self.frete_input) #VERIFICAR
        cost_group.setLayout(cost_layout)
        left_v_layout.addWidget(cost_group)

        # --- (MODIFICADO) Grupo 3: Carregar Planilha (Opcional) --- (Era 2)
        file_group = QGroupBox("3. Carregar Planilha (Opcional)")
        file_layout = QVBoxLayout()
        self.file_label = QLabel("Nenhum projeto ativo.")
        file_button_layout = QHBoxLayout()
        self.select_file_btn = QPushButton("Selecionar Planilha")
        self.import_dxf_btn = QPushButton("Importar DXF(s)") # <<< NOVO BOTÃO >>>
        self.clear_excel_btn = QPushButton("Limpar Planilha")
        file_button_layout.addWidget(self.select_file_btn)
        file_button_layout.addWidget(self.import_dxf_btn)
        file_button_layout.addWidget(self.clear_excel_btn)
        file_layout.addWidget(self.file_label)
        file_layout.addLayout(file_button_layout)
        file_group.setLayout(file_layout)
        left_v_layout.addWidget(file_group)

        # --- (MODIFICADO) Grupo 4: Informações da Peça --- (Era 3)
        manual_group = QGroupBox("4. Informações da Peça")
        manual_layout = QFormLayout()
        manual_layout.setLabelAlignment(Qt.AlignRight)
        manual_layout.setVerticalSpacing(8)
        self.projeto_input = QLineEdit()
        self.projeto_input.setReadOnly(True)
        manual_layout.addRow("Nº do Projeto Ativo:", self.projeto_input)
        self.nome_input = QLineEdit()
        self.generate_code_btn = QPushButton("Gerar Código")
        name_layout = QHBoxLayout()
        name_layout.addWidget(self.nome_input)
        name_layout.addWidget(self.generate_code_btn)
        name_layout.setSpacing(5)
        manual_layout.addRow("Nome/ID da Peça:", name_layout)
        self.forma_combo = QComboBox()
        self.forma_combo.addItems(['rectangle', 'circle', 'right_triangle', 'trapezoid', 'dxf_shape'])
        self.espessura_input, self.qtd_input = QLineEdit(), QLineEdit()
        manual_layout.addRow("Forma:", self.forma_combo)
        manual_layout.addRow("Espessura (mm):", self.espessura_input)
        manual_layout.addRow("Quantidade:", self.qtd_input)
        self.largura_input, self.altura_input = QLineEdit(), QLineEdit()
        self.diametro_input, self.rt_base_input, self.rt_height_input = QLineEdit(), QLineEdit(), QLineEdit()
        self.trapezoid_large_base_input, self.trapezoid_small_base_input, self.trapezoid_height_input = QLineEdit(), QLineEdit(), QLineEdit()
        self.largura_row = [QLabel("Largura:"), self.largura_input]; manual_layout.addRow(*self.largura_row)
        self.altura_row = [QLabel("Altura:"), self.altura_input]; manual_layout.addRow(*self.altura_row)
        self.diametro_row = [QLabel("Diâmetro:"), self.diametro_input]; manual_layout.addRow(*self.diametro_row)
        self.rt_base_row = [QLabel("Base Triângulo:"), self.rt_base_input]; manual_layout.addRow(*self.rt_base_row)
        self.rt_height_row = [QLabel("Altura Triângulo:"), self.rt_height_input]; manual_layout.addRow(*self.rt_height_row)
        self.trap_large_base_row = [QLabel("Base Maior:"), self.trapezoid_large_base_input]; manual_layout.addRow(*self.trap_large_base_row)
        self.trap_small_base_row = [QLabel("Base Menor:"), self.trapezoid_small_base_input]; manual_layout.addRow(*self.trap_small_base_row)
        self.trap_height_row = [QLabel("Altura:"), self.trapezoid_height_input]; manual_layout.addRow(*self.trap_height_row)
        manual_group.setLayout(manual_layout)
        left_v_layout.addWidget(manual_group)
        left_v_layout.addStretch()
        
        top_h_layout.addWidget(left_panel_widget) # Adiciona o painel esquerdo ao layout horizontal

        # --- (MODIFICADO) Grupo 5: Furos (Painel Direito) --- (Era 4)
        furos_main_group = QGroupBox("5. Adicionar Furos")
        furos_main_layout = QVBoxLayout()
        self.rep_group = QGroupBox("Furação Rápida")
        rep_layout = QFormLayout()
        self.rep_diam_input, self.rep_offset_input = QLineEdit(), QLineEdit()
        rep_layout.addRow("Diâmetro Furos:", self.rep_diam_input)
        rep_layout.addRow("Offset Borda:", self.rep_offset_input)
        self.replicate_btn = QPushButton("Replicar Furos")
        rep_layout.addRow(self.replicate_btn)
        self.rep_group.setLayout(rep_layout)
        furos_main_layout.addWidget(self.rep_group)
        man_group = QGroupBox("Furos Manuais")
        man_layout = QVBoxLayout()
        man_form_layout = QFormLayout()
        self.diametro_furo_input, self.pos_x_input, self.pos_y_input = QLineEdit(), QLineEdit(), QLineEdit()
        man_form_layout.addRow("Diâmetro:", self.diametro_furo_input)
        man_form_layout.addRow("Posição X:", self.pos_x_input)
        man_form_layout.addRow("Posição Y:", self.pos_y_input)
        self.add_furo_btn = QPushButton("Adicionar Furo Manual")
        man_layout.addLayout(man_form_layout)
        man_layout.addWidget(self.add_furo_btn)
        self.furos_table = QTableWidget(0, 4)
        self.furos_table.setMaximumHeight(150)
        self.furos_table.setHorizontalHeaderLabels(["Diâmetro", "Pos X", "Pos Y", "Ação"])
        man_layout.addWidget(self.furos_table)
        man_group.setLayout(man_layout)
        furos_main_layout.addWidget(man_group)
        furos_main_group.setLayout(furos_main_layout)
        top_h_layout.addWidget(furos_main_group, stretch=1)

        # Container para o layout superior
        top_container_widget = QWidget()
        top_container_widget.setLayout(top_h_layout)

        # --- (MODIFICADO) Grupo 6: Lista de Peças --- (Era 5)
        list_group = QGroupBox("6. Lista de Peças para Produção")
        list_layout = QVBoxLayout()
        self.pieces_table = QTableWidget()
        self.table_headers = [col.replace('_', ' ').title() for col in self.colunas_df] + ["Ações"]
        self.pieces_table.setColumnCount(len(self.table_headers))
        self.pieces_table.setHorizontalHeaderLabels(self.table_headers)
        self.pieces_table.verticalHeader().setDefaultSectionSize(28) 
        self.pieces_table.setMinimumHeight(120)
        
        list_layout.addWidget(self.pieces_table)
        self.dir_label = QLabel("Nenhum projeto ativo. Inicie um novo projeto.")
        self.dir_label.setStyleSheet("font-style: italic; color: grey;")
        list_layout.addWidget(self.dir_label)
        process_buttons_layout = QHBoxLayout()
        self.conclude_project_btn = QPushButton("Projeto Concluído")
        self.export_excel_btn = QPushButton("Exportar para Excel")
        self.process_pdf_btn, self.process_dxf_btn, self.process_all_btn = QPushButton("Gerar PDFs"), QPushButton("Gerar DXFs"), QPushButton("Gerar PDFs e DXFs")
        process_buttons_layout.addWidget(self.export_excel_btn)
        process_buttons_layout.addWidget(self.conclude_project_btn)
        process_buttons_layout.addStretch()
        self.calculate_nesting_btn = QPushButton("Calcular Aproveitamento")
        process_buttons_layout.addWidget(self.calculate_nesting_btn)
        process_buttons_layout.addWidget(self.process_pdf_btn)
        process_buttons_layout.addWidget(self.process_dxf_btn)
        process_buttons_layout.addWidget(self.process_all_btn)
        list_layout.addLayout(process_buttons_layout)
        list_group.setLayout(list_layout)

        # --- Barra de Log/Execução ---
        log_group = QGroupBox("Log de Execução")
        log_layout = QVBoxLayout()
        self.log_text = QTextEdit()
        self.log_text.setObjectName("logExecution") # Adicionado para estilo
        log_layout.addWidget(self.log_text)
        log_group.setLayout(log_layout)
        
        # <<< MUDANÇA ESTRUTURAL 2: USO DO QSPLITTER PARA O LAYOUT VERTICAL >>>
        v_splitter = QSplitter(Qt.Vertical)
        #v_splitter.addWidget(top_container_widget)
        v_splitter.addWidget(list_group)
        v_splitter.addWidget(log_group)

        v_splitter.setStretchFactor(0, 1)
        v_splitter.setStretchFactor(1, 0)
        v_splitter.setSizes([400, 150])

        self.add_piece_btn = QPushButton("Adicionar Peça à Lista")
        main_layout.addWidget(top_container_widget)
        main_layout.addWidget(v_splitter)
        main_layout.addWidget(self.add_piece_btn)

        # --- Barra de Progresso ---
        self.progress_bar = QProgressBar()
        main_layout.addWidget(self.progress_bar)
        
        self.statusBar().showMessage("Pronto")
        
        # --- Aplicação de Estilos Específicos via objectName ---
        self.start_project_btn.setObjectName("primaryButton")
        self.conclude_project_btn.setObjectName("successButton")
        self.calculate_nesting_btn.setObjectName("warningButton")

    def connect_signals(self):
        """Método para centralizar todas as conexões de sinais e slots."""
        self.calculate_nesting_btn.clicked.connect(self.open_nesting_dialog)
        self.start_project_btn.clicked.connect(self.start_new_project)
        self.history_btn.clicked.connect(self.show_history_dialog)
        self.select_file_btn.clicked.connect(self.select_file)
        self.import_dxf_btn.clicked.connect(self.import_dxfs) # <<< CONEXÃO DO SINAL >>>
        self.clear_excel_btn.clicked.connect(self.clear_excel_data)
        self.generate_code_btn.clicked.connect(self.generate_piece_code)
        self.add_piece_btn.clicked.connect(self.add_manual_piece)
        self.forma_combo.currentTextChanged.connect(self.update_dimension_fields)
        self.replicate_btn.clicked.connect(self.replicate_holes)
        self.add_furo_btn.clicked.connect(self.add_furo_temp)
        self.process_pdf_btn.clicked.connect(self.start_pdf_generation)
        self.process_dxf_btn.clicked.connect(self.start_dxf_generation)
        self.process_all_btn.clicked.connect(self.start_all_generation)
        self.conclude_project_btn.clicked.connect(self.conclude_project)
        self.export_excel_btn.clicked.connect(self.export_project_to_excel)
    # =====================================================================
    # --- INÍCIO: NOVA FUNÇÃO PARA OFFSET DINÂMICO (COMPARTILHADA) ---
    def _get_dynamic_offset_and_margin(self, espessura, default_offset, default_margin):
        """Retorna o offset e a margem com base na espessura."""
        # --- CORREÇÃO: A função agora prioriza o input do usuário se for diferente do padrão '8'. ---
        # Se o usuário inseriu um valor diferente do padrão (8), usa o valor do usuário.
        if abs(default_offset - 8.0) > 1e-5:
            return default_offset, default_margin

        if 0 < espessura <= 6.35: return 5, 10
        elif 6.35 < espessura <= 15.88: return 10, default_margin
        elif 15.88 < espessura <= 20: return 17, default_margin
        elif abs(espessura - 22.22) < 1e-5: return 20, default_margin
        elif 25.4 <= espessura <= 38: return 25, default_margin
        return default_offset, default_margin
    # --- FIM: NOVA FUNÇÃO PARA OFFSET DINÂMICO ---

    def start_new_project(self):
        parent_dir = QFileDialog.getExistingDirectory(self, "Selecione a Pasta Principal para o Novo Projeto")
        if not parent_dir: return
        project_name, ok = QInputDialog.getText(self, "Novo Projeto", "Digite o nome ou número do novo projeto:")
        if ok and project_name:
            project_path = os.path.join(parent_dir, project_name)
            if os.path.exists(project_path):
                reply = QMessageBox.question(self, 'Diretório Existente', f"A pasta '{project_name}' já existe.\nDeseja usá-la como o diretório do projeto ativo?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                if reply == QMessageBox.No: return
            else:
                try: os.makedirs(project_path)
                except OSError as e: QMessageBox.critical(self, "Erro ao Criar Pasta", f"Não foi possível criar o diretório do projeto:\n{e}"); return
            self._clear_session(clear_project_number=True)
            self.project_directory = project_path
            self.projeto_input.setText(project_name)
            self.dir_label.setText(f"Projeto Ativo: {self.project_directory}")
            self.dir_label.setStyleSheet("font-style: normal; color: #E2E8F0;") # Cor do texto do tema
            self.log_text.append(f"\n--- NOVO PROJETO INICIADO: {project_name} ---")
            self.log_text.append(f"Arquivos serão salvos em: {self.project_directory}")
            self.set_initial_button_state()

    def set_initial_button_state(self):
        is_project_active = self.project_directory is not None
        has_items = not (self.excel_df.empty and self.manual_df.empty)
        self.calculate_nesting_btn.setEnabled(is_project_active and has_items)
        self.start_project_btn.setEnabled(True)
        self.history_btn.setEnabled(True)
        self.select_file_btn.setEnabled(is_project_active)
        self.import_dxf_btn.setEnabled(is_project_active) # <<< ATUALIZAÇÃO DE ESTADO >>>
        self.clear_excel_btn.setEnabled(is_project_active and not self.excel_df.empty)
        self.generate_code_btn.setEnabled(is_project_active)
        self.add_piece_btn.setEnabled(is_project_active)
        self.replicate_btn.setEnabled(is_project_active)
        self.add_furo_btn.setEnabled(is_project_active)
        self.process_pdf_btn.setEnabled(is_project_active and has_items)
        self.process_dxf_btn.setEnabled(is_project_active and has_items)
        self.process_all_btn.setEnabled(is_project_active and has_items)
        self.conclude_project_btn.setEnabled(is_project_active and has_items)
        self.export_excel_btn.setEnabled(is_project_active and has_items)
        self.progress_bar.setVisible(False)

    def show_history_dialog(self):
        dialog = HistoryDialog(self.history_manager, self)
        if dialog.exec_() == QDialog.Accepted:
            loaded_pieces = dialog.loaded_project_data
            if loaded_pieces:
                project_number_loaded = loaded_pieces[0].get('project_number') if loaded_pieces and 'project_number' in loaded_pieces[0] else dialog.project_list_widget.currentItem().text()
                self.start_new_project_from_history(project_number_loaded, loaded_pieces)
    
    def start_new_project_from_history(self, project_name, pieces_data):
        parent_dir = QFileDialog.getExistingDirectory(self, f"Selecione uma pasta para o projeto '{project_name}'")
        if not parent_dir: return
        project_path = os.path.join(parent_dir, project_name)
        os.makedirs(project_path, exist_ok=True)
        self._clear_session(clear_project_number=True)
        self.project_directory = project_path
        self.projeto_input.setText(project_name)
        self.excel_df = pd.DataFrame(columns=self.colunas_df)
        self.manual_df = pd.DataFrame(pieces_data)
        self.dir_label.setText(f"Projeto Ativo: {self.project_directory}"); self.dir_label.setStyleSheet("font-style: normal; color: #E2E8F0;")
        self.log_text.append(f"\n--- PROJETO DO HISTÓRICO CARREGADO: {project_name} ---")
        self.update_table_display()
        self.set_initial_button_state()

    def start_pdf_generation(self): self.start_processing(generate_pdf=True, generate_dxf=False)
    def start_dxf_generation(self): self.start_processing(generate_pdf=False, generate_dxf=True)
    def start_all_generation(self): self.start_processing(generate_pdf=True, generate_dxf=True)

    def start_processing(self, generate_pdf, generate_dxf):
        if not self.project_directory:
            QMessageBox.warning(self, "Nenhum Projeto Ativo", "Inicie um novo projeto antes de gerar arquivos."); return
        project_number = self.projeto_input.text().strip()
        if not project_number:
            QMessageBox.warning(self, "Número do Projeto Ausente", "Por favor, defina um número para o projeto ativo."); return
        # --- CORREÇÃO FUTUREWARNING: Concatena apenas os dataframes não vazios ---
        dfs_to_concat = [df for df in [self.excel_df, self.manual_df] if not df.empty]
        if not dfs_to_concat:
            QMessageBox.warning(self, "Aviso", "A lista de peças está vazia."); return
        combined_df = pd.concat(dfs_to_concat, ignore_index=True)
        # --- FIM CORREÇÃO ---
        self.set_buttons_enabled_on_process(False)
        self.progress_bar.setVisible(True); self.progress_bar.setValue(0); self.log_text.clear()
        self.process_thread = ProcessThread(combined_df.copy(), generate_pdf, generate_dxf, self.project_directory, project_number)
        self.process_thread.update_signal.connect(self.log_text.append)
        self.process_thread.progress_signal.connect(self.progress_bar.setValue)
        self.process_thread.finished_signal.connect(self.processing_finished)
        self.process_thread.start()

    def processing_finished(self, success, message):
        self.set_buttons_enabled_on_process(True); self.progress_bar.setVisible(False)
        msgBox = QMessageBox.information if success else QMessageBox.critical
        msgBox(self, "Concluído" if success else "Erro", message); self.statusBar().showMessage("Pronto")
    
    def conclude_project(self):
        project_number = self.projeto_input.text().strip()
        if not project_number:
            QMessageBox.warning(self, "Projeto sem Número", "O projeto ativo não tem um número definido.")
            return
        reply = QMessageBox.question(self, 'Concluir Projeto', f"Deseja salvar e concluir o projeto '{project_number}'?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            # --- CORREÇÃO FUTUREWARNING: Concatena apenas os dataframes não vazios ---
            dfs_to_concat = [df for df in [self.excel_df, self.manual_df] if not df.empty]
            if dfs_to_concat:
                combined_df = pd.concat(dfs_to_concat, ignore_index=True)
            # --- FIM CORREÇÃO ---
                combined_df['project_number'] = project_number
                combined_df['project_number'] = project_number
                self.history_manager.save_project(project_number, combined_df)
                self.log_text.append(f"Projeto '{project_number}' salvo no histórico.")
            self._clear_session(clear_project_number=True)
            self.project_directory = None
            self.dir_label.setText("Nenhum projeto ativo. Inicie um novo projeto."); self.dir_label.setStyleSheet("font-style: italic; color: grey;")
            self.set_initial_button_state()
            self.log_text.append(f"\n--- PROJETO '{project_number}' CONCLUÍDO ---")

    def open_nesting_dialog(self):
        # --- CORREÇÃO FUTUREWARNING: Concatena apenas os dataframes não vazios ---
        dfs_to_concat = [df for df in [self.excel_df, self.manual_df] if not df.empty]
        if not dfs_to_concat:
            QMessageBox.warning(self, "Lista Vazia", "Não há peças na lista para calcular o aproveitamento.")
            return
        combined_df = pd.concat(dfs_to_concat, ignore_index=True)
        # --- CORREÇÃO: Inclui 'circle' na verificação de formas válidas ---
        valid_df = combined_df[combined_df['forma'].isin(['rectangle', 'circle', 'right_triangle', 'trapezoid', 'dxf_shape'])].copy()
        if valid_df.empty:
            QMessageBox.information(self, "Nenhuma Peça Válida", "O cálculo de aproveitamento só pode ser feito com peças da forma 'rectangle', 'circle', 'right_triangle', 'trapezoid' ou 'dxf_shape'.")
            return
        # Passa o DataFrame com as formas válidas para o diálogo
        dialog = NestingDialog(valid_df, self)
        dialog.exec_()

    def _get_export_parameters(self):
        """Abre um diálogo para obter os parâmetros de exportação."""
        dialog = QDialog(self)
        dialog.setWindowTitle("Parâmetros de Exportação e Aproveitamento")
        layout = QFormLayout(dialog)

        # Campos de entrada
        chapa_largura_input = QLineEdit("3000")
        chapa_altura_input = QLineEdit("1500")
        offset_input = QLineEdit("8")
        margin_input = QLineEdit("10")
        method_combo = QComboBox()
        method_combo.addItems(["Plasma/Laser", "Guilhotina"])

        layout.addRow("Largura da Chapa (mm):", chapa_largura_input)
        layout.addRow("Altura da Chapa (mm):", chapa_altura_input)
        layout.addRow("Método de Corte:", method_combo)
        layout.addRow("Offset entre Peças (mm) [Plasma/Laser]:", offset_input)
        layout.addRow("Margem da Chapa (mm) [Plasma/Laser]:", margin_input)

        # Botões OK e Cancelar
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addRow(button_box)

        if dialog.exec_() == QDialog.Accepted:
            try:
                params = {
                    "chapa_largura": float(chapa_largura_input.text()),
                    "chapa_altura": float(chapa_altura_input.text()),
                    "offset": float(offset_input.text()),
                    "margin": float(margin_input.text()),
                    "method": method_combo.currentText()
                }
                return params
            except (ValueError, TypeError):
                QMessageBox.critical(self, "Erro de Entrada", "Valores de chapa, offset e margem devem ser numéricos.")
                return None
        return None

    def export_project_to_excel(self):
        params = self._get_export_parameters()
        if not params: return

        project_number = self.projeto_input.text().strip()
        if not project_number:
            QMessageBox.warning(self, "Nenhum Projeto Ativo", "Inicie um novo projeto para poder exportá-lo.")
            return

        dfs_to_concat = [df for df in [self.excel_df, self.manual_df] if not df.empty]
        if not dfs_to_concat:
            QMessageBox.warning(self, "Lista Vazia", "Não há peças na lista para exportar.")
            return
        combined_df = pd.concat(dfs_to_concat, ignore_index=True)

        default_filename = os.path.join(self.project_directory, f"CUSTO_PLASMA-LASER_V6_{project_number}.xlsx")
        save_path, _ = QFileDialog.getSaveFileName(self, "Salvar Resumo do Projeto", default_filename, "Excel Files (*.xlsx)")
        if not save_path:
            return

        self.set_buttons_enabled_on_process(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.log_text.clear()
        self.log_text.append("Iniciando exportação para Excel...")
        QApplication.processEvents()

        try:
            template_path = 'planilha-dbx.xlsx'
            if not os.path.exists(template_path):
                QMessageBox.critical(self, "Template Não Encontrado", f"O arquivo modelo '{template_path}' não foi encontrado.")
                return

            wb = load_workbook(template_path)
            ws = wb.active
            
            # --- PASSO 1: PREENCHER IMPOSTO (A2) E FRETE (C2) ---
            try:
                imposto_val = float(self.imposto_input.text().replace(',', '.'))
                frete_val = float(self.frete_input.text().replace(',', '.'))
                ws['A2'] = imposto_val
                ws['C2'] = frete_val
                self.log_text.append(f"Imposto ({imposto_val}) e Frete ({frete_val}) preenchidos nas células A2 e C2.")
            except ValueError:
                QMessageBox.warning(self, "Valor Inválido", "Valores de Imposto e Frete devem ser numéricos. Usando 0.")
                ws['A2'] = 0
                ws['C2'] = 0

            self.log_text.append("Preenchendo lista de peças...")
            QApplication.processEvents()
            
            start_row = 4 # A lista de peças sempre começará a ser preenchida na linha 4
            last_filled_row = start_row - 1 # Guarda a última linha preenchida

            # --- PASSO 2: PREENCHER A LISTA DE PEÇAS (COLUNAS A ATÉ I) ---
            for index, (_, row_data) in enumerate(combined_df.iterrows()):
                current_row = start_row + index
                last_filled_row = current_row # Atualiza a última linha
                
                ws.cell(row=current_row, column=1, value=project_number)
                ws.cell(row=current_row, column=2, value=row_data.get('nome_arquivo', ''))
                
                qtd_peca = row_data.get('qtd', 0)
                ws.cell(row=current_row, column=3, value=qtd_peca)
                
                forma = str(row_data.get('forma', '')).lower()
                largura, altura = row_data.get('largura', 0), row_data.get('altura', 0)
                forma_map = {'circle': 'C', 'trapezoid': 'TP', 'right_triangle': 'T'}
                forma_abreviada = 'Q' if forma == 'rectangle' and largura == altura and largura > 0 else forma_map.get(forma, 'R' if forma == 'rectangle' else '')
                ws.cell(row=current_row, column=4, value=forma_abreviada)

                furos = row_data.get('furos', [])
                num_furos = len(furos) if isinstance(furos, list) else 0
                ws.cell(row=current_row, column=5, value=num_furos)
                ws.cell(row=current_row, column=6, value=furos[0].get('diam', 0) if num_furos > 0 else 0)
                
                # Esta é a 'ESPESSURA' da Coluna G
                espessura_peca = row_data.get('espessura', 0)
                ws.cell(row=current_row, column=7, value=espessura_peca) # <--- COLUNA G
                
                ws.cell(row=current_row, column=8, value=largura)
                ws.cell(row=current_row, column=9, value=altura)
                
                self.progress_bar.setValue(int(((index + 1) / (len(combined_df) * 2)) * 100))

            # --- PASSO 3: OCULTAR LINHAS (DA ÚLTIMA PEÇA ATÉ A 207) ---
            # (Ajustado para 211, que é a linha "RELATÓRIO DE APROVEITAMENTO")
            try:
                start_hide_row = last_filled_row + 1
                end_hide_row = 207 # Oculta até a linha ANTES do relatório
                if start_hide_row <= end_hide_row:
                    ws.row_dimensions.group(start_hide_row, end_hide_row, hidden=True)
                    self.log_text.append(f"Linhas da {start_hide_row} até {end_hide_row} ocultadas com sucesso.")
                else:
                    self.log_text.append(f"Nenhuma linha para ocultar (Última linha: {last_filled_row}).")
            except Exception as e:
                self.log_text.append(f"AVISO: Falha ao ocultar linhas. {e}")
            
            self.log_text.append("Calculando aproveitamento de chapas...")
            QApplication.processEvents()

            valid_nesting_df = combined_df[combined_df['forma'].isin(['rectangle', 'circle', 'right_triangle', 'trapezoid', 'dxf_shape'])].copy()
            valid_nesting_df['espessura'] = valid_nesting_df['espessura'].astype(float)
            
            # Agrupa as peças pela espessura (Coluna G)
            grouped = valid_nesting_df.groupby('espessura') # <--- AGRUPANDO PELA COLUNA G
            
            current_row = 212 # O relatório sempre começará na linha 212
            ws.cell(row=current_row, column=1, value="RELATÓRIO DE APROVEITAMENTO DE CHAPA").font = Font(bold=True, size=14)
            current_row += 2

            # --- (NOVO) Variáveis para Perca Ponderada REAL ---
            total_perca_ponderada_real = 0.0
            total_pecas_contadas_real = 0.0
            perda_results_map = {} # Dicionário para salvar {espessura: percent_loss}

            # --- Inicia o loop para processar por espessura (baseado na Coluna G) ---
            for espessura, group in grouped:
                is_guillotine = params["method"] == "Guilhotina"
                
                if is_guillotine:
                    # ... (lógica da guilhotina) ...
                    current_offset, refila = 0, 2 * espessura
                    sheet_width_for_calc, sheet_height_for_calc = params["chapa_largura"] - refila, params["chapa_altura"]
                    effective_margin = 0
                else: # Plasma/Laser
                    # ... (lógica plasma/laser) ...
                    current_offset, _ = self._get_dynamic_offset_and_margin(espessura, params["offset"], params["margin"])
                    effective_margin = 10 - (current_offset / 2)
                    sheet_width_for_calc, sheet_height_for_calc = params["chapa_largura"], params["chapa_altura"]

                pecas_para_calcular = []
                total_pecas_neste_grupo = 0

                for _, row in group.iterrows():
                    qtd = int(row['qtd'])
                    total_pecas_neste_grupo += qtd
                    # ... (adiciona peças ao 'pecas_para_calcular') ...
                    if row['forma'] == 'rectangle' and row['largura'] > 0 and row['altura'] > 0:
                        pecas_para_calcular.append({'forma': 'rectangle', 'largura': row['largura'], 'altura': row['altura'], 'quantidade': qtd})
                    elif row['forma'] == 'circle' and row['diametro'] > 0:
                        pecas_para_calcular.append({'forma': 'circle', 'largura': row['diametro'], 'altura': row['diametro'], 'diametro': row['diametro'], 'quantidade': qtd})
                    elif row['forma'] == 'right_triangle' and row['rt_base'] > 0 and row['rt_height'] > 0:
                        pecas_para_calcular.append({'forma': 'right_triangle', 'largura': row['rt_base'], 'altura': row['rt_height'], 'quantidade': qtd})
                    elif row['forma'] == 'trapezoid' and row['trapezoid_large_base'] > 0 and row['trapezoid_height'] > 0:
                        pecas_para_calcular.append({'forma': 'trapezoid', 'largura': row['trapezoid_large_base'], 'altura': row['trapezoid_height'], 'small_base': row['trapezoid_small_base'], 'quantidade': qtd})
                    elif row['forma'] == 'dxf_shape' and row['largura'] > 0 and row['altura'] > 0:
                        pecas_para_calcular.append({'forma': 'dxf_shape', 'largura': row['largura'], 'altura': row['altura'], 'dxf_path': row['dxf_path'], 'quantidade': qtd})

                if not pecas_para_calcular: continue

                self.log_text.append(f"Otimizando espessura {espessura}mm (pode levar um momento)...")
                QApplication.processEvents()
                
                pecas_com_offset = []
                for p in pecas_para_calcular:
                    p_copy = p.copy()
                    p_copy['largura'] += current_offset
                    p_copy['altura'] += current_offset
                    if 'small_base' in p_copy: p_copy['small_base'] += current_offset
                    pecas_com_offset.append(p_copy)
                
                # --- O CÁLCULO DE NESTING É FEITO AQUI ---
                resultado = orquestrar_planos_de_corte(sheet_width_for_calc, sheet_height_for_calc, pecas_com_offset, current_offset, effective_margin, espessura, is_guillotine, status_signal_emitter=None)
                
                if not resultado: continue

                # --- (NOVO) PASSO 4: Captura da PERCA REAL calculada ---
                percentual_perda = resultado.get('percentual_perda_total_sucata', 0)
                
                # Salva o resultado (ex: {12.7: 20.43})
                perda_results_map[espessura] = percentual_perda 
                
                total_perca_ponderada_real += (percentual_perda * total_pecas_neste_grupo)
                total_pecas_contadas_real += total_pecas_neste_grupo

                # --- ESCRITA DO RELATÓRIO DE NESTING (A212 em diante) ---
                ws.cell(row=current_row, column=1, value=f"Espessura: {espessura} mm").font = Font(bold=True, size=12)
                current_row += 1
                total_chapas_usadas = resultado['total_chapas']
                peso_total_chapas_kg = (params["chapa_largura"]/1000) * (params["chapa_altura"]/1000) * espessura * 7.85 * total_chapas_usadas
                ws.cell(row=current_row, column=1, value=f"Total de Chapas: {total_chapas_usadas}")
                ws.cell(row=current_row, column=2, value=f"Aproveitamento: {resultado['aproveitamento_geral']}")
                ws.cell(row=current_row, column=3, value=f"Peso Total das Chapas: {peso_total_chapas_kg:.2f} kg").font = Font(bold=True)
                current_row += 2

                for i, plano_info in enumerate(resultado['planos_unicos']):
                    # ... (escreve planos de corte) ...
                    ws.cell(row=current_row, column=1, value=f"Plano de Corte {i+1} (Repetir {plano_info['repeticoes']}x)").font = Font(italic=True)
                    current_row += 1
                    ws.cell(row=current_row, column=2, value="Peças neste plano:")
                    current_row += 1
                    for item in plano_info['resumo_pecas']:
                        ws.cell(row=current_row, column=3, value=f"- {item['qtd']}x de {item['tipo']}")
                        current_row += 1
                    current_row += 1

                sucata_info = resultado.get('sucata_detalhada')
                if sucata_info:
                    # ... (escreve todo o detalhamento de sucata) ...
                    bold_font = Font(bold=True)
                    ws.cell(row=current_row, column=1, value="Peso do Offset (perda de corte):").font = bold_font
                    ws.cell(row=current_row, column=2, value=f"{sucata_info['peso_offset']:.2f} kg")
                    current_row += 2
                    ws.cell(row=current_row, column=1, value="Sobras Aproveitáveis (Retalhos > 300x300 mm)").font = bold_font
                    current_row += 1
                    sobras_aproveitaveis = sucata_info['sobras_aproveitaveis']
                    if not sobras_aproveitaveis:
                        ws.cell(row=current_row, column=2, value="- Nenhuma")
                        current_row += 1
                    else:
                        from collections import Counter
                        contagem = Counter((s['largura'], s['altura'], f"{s['peso']:.2f}") for s in sobras_aproveitaveis for _ in range(s['quantidade']))
                        total_peso_aproveitavel = sum(s['peso'] * s['quantidade'] for s in sobras_aproveitaveis)
                        for (larg, alt, peso_unit), qtd in contagem.items():
                            ws.cell(row=current_row, column=2, value=f"- {qtd}x de {larg:.0f}x{alt:.0f} mm (Peso unit: {peso_unit} kg)")
                            current_row += 1
                        ws.cell(row=current_row, column=2, value=f"Peso Total Aproveitável: {total_peso_aproveitavel:.2f} kg").font = bold_font
                        current_row += 1
                    current_row += 1
                    ws.cell(row=current_row, column=1, value="Sucatas com Dimensões").font = bold_font
                    current_row += 1
                    sucatas_dim = sucata_info['sucatas_dimensionadas']
                    if not sucatas_dim:
                        ws.cell(row=current_row, column=2, value="- Nenhuma")
                        current_row += 1
                    else:
                        from collections import Counter
                        contagem = Counter((s['largura'], s['altura'], f"{s['peso']:.2f}") for s in sucatas_dim for _ in range(s['quantidade']))
                        total_peso_sucata_dim = sum(s['peso'] * s['quantidade'] for s in sucatas_dim)
                        for (larg, alt, peso_unit), qtd in contagem.items():
                            ws.cell(row=current_row, column=2, value=f"- {qtd}x de {larg:.0f}x{alt:.0f} mm (Peso unit: {peso_unit} kg)")
                            current_row += 1
                        ws.cell(row=current_row, column=2, value=f"Peso Total (Sucata Dimensionada): {total_peso_sucata_dim:.2f} kg").font = bold_font
                        current_row += 1
                    current_row += 1
                    ws.cell(row=current_row, column=1, value="Demais Sucatas (cavacos, etc):").font = bold_font
                    ws.cell(row=current_row, column=2, value=f"{sucata_info['peso_demais_sucatas']:.2f} kg")
                    current_row += 2
                    ws.cell(row=current_row, column=1, value="Resumo da Perda Total (Sucata + Processo + Offset):").font = bold_font
                    ws.cell(row=current_row, column=2, value=f"{resultado.get('peso_perda_total_sucata', 0):.2f} kg")
                    ws.cell(row=current_row, column=3, value=f"({resultado.get('percentual_perda_total_sucata', 0):.2f} % do total)").font = Font(italic=True)
                    current_row += 2

                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
                cell = ws.cell(row=current_row, column=1)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                current_row += 2
                self.progress_bar.setValue(50 + int((current_row / 400) * 50))
            
            # --- FIM DO LOOP DE NESTING ---

            # --- (NOVO) PASSO 5: PREENCHER PERDA (D2) COM BASE NOS RESULTADOS REAIS ---
            if total_pecas_contadas_real > 0:
                avg_loss_real = total_perca_ponderada_real / total_pecas_contadas_real
                # A perca é um percentual (ex: 20.01), mas a célula D2 espera um decimal (ex: 0.2001)
                ws['D2'] = avg_loss_real / 100.0 
                self.log_text.append(f"Perca média ponderada REAL ({avg_loss_real:.2f}%) preenchida em D2.")
            else:
                ws['D2'] = 0
                self.log_text.append("Nenhuma peça para calcular perca real. Preenchido 0 em D2.")

            # --- (CORRIGIDO) PASSO 6: PREENCHER COLUNA W CONFORME SUA SOLICITAÇÃO ---
            self.log_text.append("Atualizando tabela de perdas (Coluna W) com resultados do nesting...")
            
            # (CORRIGIDO) Loop para encontrar a espessura na Coluna V, começando da linha 215
            # A sua tabela vai da linha 215 (Índice 1) até 239 (Índice 25)
            for row_idx in range(215, 240): 
                esp_cell = ws.cell(row=row_idx, column=22) # Coluna V (ESPESSURA)
                if esp_cell.value is None:
                    break # Para se a tabela terminar
                try:
                    esp_template = float(esp_cell.value)
                    
                    # Verifica se temos um resultado de nesting para essa espessura
                    # ex: O mapa tem {12.7: 20.43}
                    if esp_template in perda_results_map:
                        
                        # Encontrou a espessura! (ex: 12,70 na linha 233)
                        
                        # Converte a perca (ex: 20.43) para decimal (ex: 0.2043) para o Excel
                        perda_para_escrever = perda_results_map[esp_template] / 100.0
                        
                        # Escreve a perca calculada na Coluna W (ex: 0.2043 em W233)
                        ws.cell(row=row_idx, column=23, value=perda_para_escrever) # <--- COLUNA W
                        
                except (ValueError, TypeError):
                    continue # Ignora linhas mal formatadas na tabela

            self.log_text.append("Salvando arquivo Excel...")
            QApplication.processEvents()
            wb.save(save_path)
            self.progress_bar.setValue(100)
            self.log_text.append(f"Resumo do projeto salvo com sucesso em: {save_path}")
            QMessageBox.information(self, "Sucesso", f"O arquivo Excel foi salvo com sucesso em:\n{save_path}")
            self._generate_pdf_from_excel(save_path, len(combined_df))
        except Exception as e:
            self.log_text.append(f"ERRO ao exportar para Excel: {e}")
            QMessageBox.critical(self, "Erro na Exportação", f"Ocorreu um erro ao salvar o arquivo:\n{e}")
        finally:
            self.set_buttons_enabled_on_process(True)
            self.progress_bar.setVisible(False)
        
    # ... (Aqui termina sua função export_project_to_excel) ...
    # ... (Cole a nova função abaixo) ...

    def _generate_pdf_from_excel(self, excel_path, num_pecas):
        """
        (NOVA FUNÇÃO)
        Usa pywin32 para abrir o Excel salvo e exportá-lo como PDF.
        Esta lógica foi portada do seu script CLI.
        """
        if not PYWIN32_DISPONIVEL:
            self.log_text.append("\n[AVISO] Geração de PDF pulada. Biblioteca 'pywin32' não encontrada.")
            return

        self.log_text.append("\nIniciando geração de PDF do orçamento...")
        
        pdf_filename = os.path.splitext(excel_path)[0] + ".pdf"
        HEADER_ROW = 3 
        START_ROW = 4 
        TOTAL_ROW = 209 # Linha "TOTAIS" ou "OBS SOBRE A PRODUÇÃO"
        LAST_EMPTY_ROW = 207 

        excel = None
        workbook = None

        try:
            last_data_row = START_ROW + num_pecas - 1
            first_empty_row = last_data_row + 1
            
            range_to_hide = None
            if first_empty_row <= LAST_EMPTY_ROW:
                range_to_hide = f"{first_empty_row}:{LAST_EMPTY_ROW}"
            
            pythoncom.CoInitialize()
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            
            full_excel_path = os.path.abspath(excel_path)
            
            workbook = excel.Workbooks.Open(full_excel_path)
            
            # --- (CORREÇÃO 1) ---
            # Seleciona a planilha pelo nome exato (baseado nas suas imagens)
            # É mais seguro que usar "ActiveSheet"
            sheet = workbook.Worksheets("PEÇAS PADRÃO")
            sheet.Activate()
            # --- (FIM DA CORREÇÃO 1) ---
            
            sheet.Rows.Hidden = False # Desoculta tudo primeiro
            if range_to_hide:
                self.log_text.append(f"Ocultando linhas {range_to_hide} para o PDF...")
                sheet.Rows(range_to_hide).Hidden = True
            
            # (CORRIGIDO) Define a área de impressão para A3:V208
            # A linha 208 contém os "TOTAIS" e queremos incluí-la
            print_area_range = f"A{HEADER_ROW}:V{TOTAL_ROW}" 
            sheet.PageSetup.PrintArea = print_area_range
            
            sheet.PageSetup.Zoom = False
            sheet.PageSetup.FitToPagesWide = 1 
            sheet.PageSetup.FitToPagesTall = 1 
            sheet.PageSetup.Orientation = 2 # 2 = xlLandscape (Paisagem)
            
            full_pdf_path = os.path.abspath(pdf_filename)
            self.log_text.append(f"Exportando PDF para: {full_pdf_path}...")
            
            # Exporta APENAS a planilha ativa
            sheet.ExportAsFixedFormat(0, full_pdf_path) # 0 = xlTypePDF
            
            self.log_text.append(f"✅ SUCESSO! PDF do orçamento gerado.")

        except Exception as e:
            self.log_text.append(f"\n[ERRO] Falha ao gerar o PDF do orçamento: {e}")
            self.log_text.append("Verifique se o Excel está instalado e se o pywin32 foi registrado (pywin32_postinstall.py -install).")
        
        finally:
            # --- (CORREÇÃO 2) ---
            # Adiciona try/except para evitar o erro "Objeto desconectado"
            try:
                if workbook:
                    workbook.Close(SaveChanges=False) # Não salvar (já foi salvo pelo openpyxl)
                if excel:
                    excel.Quit()
                del excel 
                pythoncom.CoUninitialize()
            except Exception as e_cleanup:
                self.log_text.append(f"[AVISO] Erro durante a limpeza do COM: {e_cleanup}")
                try:
                    # Garante que o COM seja liberado mesmo se o excel.Quit() falhar
                    pythoncom.CoUninitialize()
                except:
                    pass 
            # --- (FIM DA CORREÇÃO 2) ---

    def _clear_session(self, clear_project_number=False):
        fields_to_clear = [self.nome_input, self.espessura_input, self.qtd_input, self.largura_input, self.altura_input, self.diametro_input, self.rt_base_input, self.rt_height_input, self.trapezoid_large_base_input, self.trapezoid_small_base_input, self.trapezoid_height_input, self.rep_diam_input, self.rep_offset_input, self.diametro_furo_input, self.pos_x_input, self.pos_y_input]
        if clear_project_number:
            fields_to_clear.append(self.projeto_input)
        for field in fields_to_clear:
            field.clear()
        self.furos_atuais = []
        self.update_furos_table()
        self.file_label.setText("Nenhum projeto ativo.")
        if clear_project_number: 
            self.excel_df = pd.DataFrame(columns=self.colunas_df)
            self.manual_df = pd.DataFrame(columns=self.colunas_df)
            self.update_table_display()

    def set_buttons_enabled_on_process(self, enabled):
        is_project_active = self.project_directory is not None
        has_items = not (self.excel_df.empty and self.manual_df.empty)
        self.calculate_nesting_btn.setEnabled(enabled and is_project_active and has_items)
        self.start_project_btn.setEnabled(enabled)
        self.history_btn.setEnabled(enabled)
        self.select_file_btn.setEnabled(enabled and is_project_active)
        self.import_dxf_btn.setEnabled(enabled and is_project_active) # <<< ATUALIZAÇÃO DE ESTADO >>>
        self.clear_excel_btn.setEnabled(enabled and is_project_active and not self.excel_df.empty)
        self.generate_code_btn.setEnabled(enabled and is_project_active)
        self.add_piece_btn.setEnabled(enabled and is_project_active)
        self.replicate_btn.setEnabled(enabled and is_project_active)
        self.add_furo_btn.setEnabled(enabled and is_project_active)
        self.process_pdf_btn.setEnabled(enabled and is_project_active and has_items)
        self.process_dxf_btn.setEnabled(enabled and is_project_active and has_items)
        self.process_all_btn.setEnabled(enabled and is_project_active and has_items)
        self.conclude_project_btn.setEnabled(enabled and is_project_active and has_items)
        self.export_excel_btn.setEnabled(enabled and is_project_active and has_items)

    def update_table_display(self):
        self.set_initial_button_state()
        # --- CORREÇÃO FUTUREWARNING: Concatena apenas os dataframes não vazios ---
        dfs_to_concat = [df for df in [self.excel_df, self.manual_df] if not df.empty]
        if dfs_to_concat:
            combined_df = pd.concat(dfs_to_concat, ignore_index=True)
        else:
            combined_df = pd.DataFrame(columns=self.colunas_df)
        self.pieces_table.blockSignals(True)
        self.pieces_table.setRowCount(0)
        self.pieces_table.blockSignals(False)

        if combined_df.empty:
            return

        self.pieces_table.setRowCount(len(combined_df))
        self.pieces_table.verticalHeader().setDefaultSectionSize(40)
        
        for i, row in combined_df.iterrows():
            for j, col in enumerate(self.colunas_df):
                value = row.get(col)
                if col == 'furos' and isinstance(value, list):
                    display_value = f"{len(value)} Furo(s)"
                elif pd.isna(value) or value == 0:
                    display_value = '-'
                else:
                    display_value = str(value)
                item = QTableWidgetItem(display_value)
                item.setTextAlignment(Qt.AlignVCenter | Qt.AlignLeft)
                self.pieces_table.setItem(i, j, item)

            action_widget = QWidget()
            action_layout = QHBoxLayout(action_widget)
            action_layout.setContentsMargins(5, 0, 5, 0)
            action_layout.setSpacing(5)
            edit_btn, delete_btn = QPushButton("Editar"), QPushButton("Excluir")
            edit_btn.clicked.connect(lambda _, r=i: self.edit_row(r))
            delete_btn.clicked.connect(lambda _, r=i: self.delete_row(r))
            action_layout.addWidget(edit_btn)
            action_layout.addWidget(delete_btn)
            self.pieces_table.setCellWidget(i, len(self.colunas_df), action_widget)

        header = self.pieces_table.horizontalHeader()
        header_map = {self.table_headers[i]: i for i in range(len(self.table_headers))}

        for col_name in ['Forma', 'Espessura', 'Qtd', 'Furos']:
            if col_name in header_map:
                header.setSectionResizeMode(header_map[col_name], QHeaderView.ResizeToContents)
        
        if 'Nome Arquivo' in header_map:
            header.setSectionResizeMode(header_map['Nome Arquivo'], QHeaderView.Stretch)
            
        dim_cols = ['Largura', 'Altura', 'Diametro', 'Rt Base', 'Rt Height', 
                    'Trapezoid Large Base', 'Trapezoid Small Base', 'Trapezoid Height']
        for col_name in dim_cols:
            if col_name in header_map:
                header.setSectionResizeMode(header_map[col_name], QHeaderView.ResizeToContents)

        if 'Ações' in header_map:
            header.setSectionResizeMode(header_map['Ações'], QHeaderView.ResizeToContents)

    def edit_row(self, row_index):
        len_excel = len(self.excel_df)
        is_from_excel = row_index < len_excel
        df_source = self.excel_df if is_from_excel else self.manual_df
        local_index = row_index if is_from_excel else row_index - len_excel
        if local_index >= len(df_source): return # Proteção contra índice inválido
        piece_data = df_source.iloc[local_index]
        self.nome_input.setText(str(piece_data.get('nome_arquivo', '')))
        self.espessura_input.setText(str(piece_data.get('espessura', '')))
        self.qtd_input.setText(str(piece_data.get('qtd', '')))
        shape = piece_data.get('forma', '')
        index = self.forma_combo.findText(shape, Qt.MatchFixedString)
        if index >= 0: self.forma_combo.setCurrentIndex(index)
        self.largura_input.setText(str(piece_data.get('largura', '')))
        self.altura_input.setText(str(piece_data.get('altura', '')))
        self.diametro_input.setText(str(piece_data.get('diametro', '')))
        self.rt_base_input.setText(str(piece_data.get('rt_base', '')))
        self.rt_height_input.setText(str(piece_data.get('rt_height', '')))
        self.trapezoid_large_base_input.setText(str(piece_data.get('trapezoid_large_base', '')))
        self.trapezoid_small_base_input.setText(str(piece_data.get('trapezoid_small_base', '')))
        self.trapezoid_height_input.setText(str(piece_data.get('trapezoid_height', '')))
        self.furos_atuais = piece_data.get('furos', []).copy() if isinstance(piece_data.get('furos'), list) else []
        self.update_furos_table()
        df_source.drop(df_source.index[local_index], inplace=True)
        df_source.reset_index(drop=True, inplace=True)
        self.log_text.append(f"Peça '{piece_data['nome_arquivo']}' carregada para edição.")
        self.update_table_display()
    
    def delete_row(self, row_index):
        len_excel = len(self.excel_df)
        is_from_excel = row_index < len_excel
        df_source = self.excel_df if is_from_excel else self.manual_df
        local_index = row_index if is_from_excel else row_index - len_excel
        if local_index >= len(df_source): return # Proteção contra índice inválido
        piece_name = df_source.iloc[local_index]['nome_arquivo']
        df_source.drop(df_source.index[local_index], inplace=True)
        df_source.reset_index(drop=True, inplace=True)
        self.log_text.append(f"Peça '{piece_name}' removida.")
        self.update_table_display()
    
    def generate_piece_code(self):
        project_number = self.projeto_input.text().strip()
        if not project_number: QMessageBox.warning(self, "Campo Obrigatório", "Inicie um projeto para definir o 'Nº do Projeto'."); return
        new_code = self.code_generator.generate_new_code(project_number, prefix='DES') #SUFIXO DOS CÓDIGOS
        if new_code: self.nome_input.setText(new_code); self.log_text.append(f"Código '{new_code}' gerado para o projeto '{project_number}'.")
    
    def add_manual_piece(self):
        try:
            nome = self.nome_input.text().strip()
            if not nome: 
                QMessageBox.warning(self, "Campo Obrigatório", "'Nome/ID da Peça' é obrigatório.")
                return

            new_piece = {'furos': self.furos_atuais.copy()}
            
            # Garante que todas as colunas do DF existam no dicionário,
            # mesmo que com valor 0 ou None, para o .loc funcionar
            for col in self.colunas_df:
                if col not in new_piece:
                    new_piece[col] = 0.0 # ou pd.NA se preferir

            # Preenche o dicionário 'new_piece' com os dados da interface
            new_piece.update({
                'nome_arquivo': nome, 
                'forma': self.forma_combo.currentText()
            })
            
            fields_map = { 
                'espessura': self.espessura_input, 
                'qtd': self.qtd_input, 
                'largura': self.largura_input, 
                'altura': self.altura_input, 
                'diametro': self.diametro_input, 
                'rt_base': self.rt_base_input, 
                'rt_height': self.rt_height_input, 
                'trapezoid_large_base': self.trapezoid_large_base_input, 
                'trapezoid_small_base': self.trapezoid_small_base_input, 
                'trapezoid_height': self.trapezoid_height_input 
            }
            
            for key, field in fields_map.items():
                new_piece[key] = float(field.text().replace(',', '.')) if field.text() else 0.0

            # --- CORREÇÃO AQUI ---
            # Em vez de pd.concat, usamos .loc para adicionar a nova linha (dicionário)
            # ao final do DataFrame. É mais limpo e evita o FutureWarning.
            self.manual_df.loc[len(self.manual_df)] = new_piece
            # --- FIM DA CORREÇÃO ---

            self.log_text.append(f"Peça '{nome}' adicionada/atualizada.")
            self._clear_session(clear_project_number=False)
            self.update_table_display()
            
        except ValueError: 
            QMessageBox.critical(self, "Erro de Valor", "Campos numéricos devem conter números válidos.")
        except Exception as e:
            QMessageBox.critical(self, "Erro Inesperado", f"Ocorreu um erro ao adicionar a peça: {e}")
    
    def select_file(self):
        if not self.project_directory:
            QMessageBox.warning(self, "Nenhum Projeto Ativo", "Inicie um projeto antes de carregar uma planilha.")
            return
        
        file_path, _ = QFileDialog.getOpenFileName(self, "Selecionar Planilha", "", "Excel Files (*.xlsx *.xls)")
        if not file_path:
            return

        try:
            df = pd.read_excel(file_path, header=0, decimal=',')
            df.columns = df.columns.str.strip().str.lower()
            self.log_text.append(f"Lendo arquivo: {os.path.basename(file_path)}")

            # --- INÍCIO DA CORREÇÃO (Leitura de Furos) ---
            # O código anterior procurava uma coluna 'furos' com JSON.
            # Esta nova lógica lê as colunas 'furo_N_...' (formato wide)
            # e as transforma no formato de lista que o app espera.

            # 1. Padronizar nomes de colunas (ex: 'furo_1_diam' -> 'furo_1_diametro')
            rename_map = {}
            for col in df.columns:
                if 'furo_' in col and col.endswith('_diam'):
                    rename_map[col] = col.replace('_diam', '_diametro')
            if rename_map:
                df = df.rename(columns=rename_map)
                self.log_text.append(f"Colunas de diâmetro padronizadas.")

            # 2. Definir os grupos de furos que vamos procurar
            #    (Aumente o 'max_furos' se houver mais grupos)
            
            # <<< ATUALIZAÇÃO: Alterado de 6 para 8 >>>
            max_furos = 8 
            
            furo_grupos = []
            for i in range(1, max_furos + 1):
                furo_grupos.append({
                    'diam': f'furo_{i}_diametro',
                    'x': f'furo_{i}_x',
                    'y': f'furo_{i}_y'
                })

            # 3. Função para ser aplicada em cada *linha* do DataFrame
            def processar_furos_da_linha(row):
                furos_encontrados = []
                for grupo in furo_grupos:
                    col_diam = grupo['diam']
                    col_x = grupo['x']
                    col_y = grupo['y']
                    
                    # Verifica se as colunas necessárias existem no DataFrame
                    if col_diam in row and col_x in row and col_y in row:
                        try:
                            # Tenta converter os valores para numérico
                            diam = pd.to_numeric(row[col_diam], errors='coerce')
                            x = pd.to_numeric(row[col_x], errors='coerce')
                            y = pd.to_numeric(row[col_y], errors='coerce')
                            
                            # Apenas adiciona se o diâmetro for um número válido e > 0
                            # e as coordenadas X/Y também forem válidas
                            if pd.notna(diam) and diam > 0 and pd.notna(x) and pd.notna(y):
                                furos_encontrados.append({
                                    'diam': float(diam),
                                    'x': float(x),
                                    'y': float(y)
                                })
                        except Exception:
                            # Ignora erros de conversão nesta linha/coluna específica
                            pass
                return furos_encontrados

            # 4. Aplica a função e cria a coluna 'furos' que o app espera
            df['furos'] = df.apply(processar_furos_da_linha, axis=1)
            
            # --- FIM DA CORREÇÃO ---

            # O resto da lógica (criar colunas faltantes, converter numéricos)
            # agora pode ser executado.
            
            df = df.loc[:, ~df.columns.duplicated()] # Remove colunas duplicadas
            
            # Garante que todas as colunas que o app precisa existam
            for col in self.colunas_df:
                if col not in df.columns: 
                    df[col] = pd.NA
            
            # Converte colunas numéricas principais (muitas colunas de furo já são numéricas)
            numeric_cols = [
                'espessura', 'qtd', 'largura', 'altura', 'diametro', 'rt_base', 'rt_height', 
                'trapezoid_large_base', 'trapezoid_small_base', 'trapezoid_height'
            ]
            
            for col in numeric_cols:
                if col in df.columns:
                    # Garante que a coluna 'furos' não seja tocada aqui
                    if col != 'furos':
                        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
            
            # Filtra o DF final para ter apenas as colunas esperadas
            self.excel_df = df[self.colunas_df].copy()
            
            self.file_label.setText(f"Planilha: {os.path.basename(file_path)}")
            self.update_table_display()
            self.log_text.append(f"Planilha '{os.path.basename(file_path)}' carregada com sucesso.")
            self.log_text.append(f"Furos processados (até {max_furos} grupos) a partir das colunas 'furo_N_...'.")

        except Exception as e:
            QMessageBox.critical(self, "Erro de Leitura", f"Falha ao ler o arquivo: {e}\n\nVerifique o console para mais detalhes.")
            print(f"Erro detalhado ao ler Excel: {e}")
            import traceback
            traceback.print_exc()
    
    def clear_excel_data(self):
        self.excel_df = pd.DataFrame(columns=self.colunas_df); self.file_label.setText("Nenhuma planilha selecionada"); self.update_table_display()

    def import_dxfs(self):
        if not self.project_directory:
            QMessageBox.warning(self, "Nenhum Projeto Ativo", "Inicie um projeto antes de importar arquivos DXF.")
            return

        file_paths, _ = QFileDialog.getOpenFileNames(self, "Selecionar Arquivos DXF", "", "DXF Files (*.dxf)")
        if not file_paths:
            return

        imported_count = 0
        for file_path in file_paths:
            largura, altura = get_dxf_bounding_box(file_path)

            if largura is not None and altura is not None:
                nome_arquivo = os.path.splitext(os.path.basename(file_path))[0]
                
                new_piece = { # type: ignore
                    'nome_arquivo': nome_arquivo,
                    'forma': 'rectangle', # Sempre será retângulo
                    'forma': 'dxf_shape',
                    'espessura': 0.0, # Padrão, para ser editado pelo usuário
                    'qtd': 1, # Padrão
                    'largura': round(largura, 2),
                    'altura': round(altura, 2),
                    'diametro': 0.0, 'rt_base': 0.0, 'rt_height': 0.0,
                    'trapezoid_large_base': 0.0, 'trapezoid_small_base': 0.0, 'trapezoid_height': 0.0,
                    'furos': [],
                    'dxf_path': file_path # Armazena o caminho do arquivo
                }
                self.manual_df = pd.concat([self.manual_df, pd.DataFrame([new_piece])], ignore_index=True)
                imported_count += 1
            else:
                self.log_text.append(f"AVISO: Não foi possível obter as dimensões do arquivo '{os.path.basename(file_path)}'. Pode estar vazio ou corrompido.")
        
        self.log_text.append(f"--- {imported_count} arquivo(s) DXF importado(s) com sucesso. ---")

    
    def replicate_holes(self):
        try:
            if self.forma_combo.currentText() != 'rectangle': QMessageBox.warning(self, "Função Indisponível", "Replicação disponível apenas para Retângulos."); return
            largura, altura = float(self.largura_input.text().replace(',', '.')), float(self.altura_input.text().replace(',', '.'))
            diam, offset = float(self.rep_diam_input.text().replace(',', '.')), float(self.rep_offset_input.text().replace(',', '.'))
            if (offset * 2) >= largura or (offset * 2) >= altura: QMessageBox.warning(self, "Offset Inválido", "Offset excede as dimensões da peça."); return
            furos = [{'diam': diam, 'x': offset, 'y': offset}, {'diam': diam, 'x': largura - offset, 'y': offset}, {'diam': diam, 'x': largura - offset, 'y': altura - offset}, {'diam': diam, 'x': offset, 'y': altura - offset}]
            self.furos_atuais.extend(furos); self.update_furos_table()
        except ValueError: QMessageBox.critical(self, "Erro de Valor", "Largura, Altura, Diâmetro e Offset devem ser números válidos.")
    
    def update_dimension_fields(self, shape):
        shape = shape.lower()
        is_rect, is_circ, is_tri, is_trap = shape == 'rectangle', shape == 'circle', shape == 'right_triangle', shape == 'trapezoid'
        for w in self.largura_row + self.altura_row: w.setVisible(is_rect)
        for w in self.diametro_row: w.setVisible(is_circ)
        for w in self.rt_base_row + self.rt_height_row: w.setVisible(is_tri)
        for w in self.trap_large_base_row + self.trap_small_base_row + self.trap_height_row: w.setVisible(is_trap)
        self.rep_group.setEnabled(is_rect)
    
    def add_furo_temp(self):
        try:
            diam, pos_x, pos_y = float(self.diametro_furo_input.text().replace(',', '.')), float(self.pos_x_input.text().replace(',', '.')), float(self.pos_y_input.text().replace(',', '.'))
            if diam <= 0: QMessageBox.warning(self, "Valor Inválido", "Diâmetro do furo deve ser maior que zero."); return
            self.furos_atuais.append({'diam': diam, 'x': pos_x, 'y': pos_y}); self.update_furos_table()
            for field in [self.diametro_furo_input, self.pos_x_input, self.pos_y_input]: field.clear()
        except ValueError: QMessageBox.critical(self, "Erro de Valor", "Campos de furo devem ser números válidos.")
    
    def update_furos_table(self):
        self.furos_table.setRowCount(0); self.furos_table.setRowCount(len(self.furos_atuais))
        for i, furo in enumerate(self.furos_atuais):
            self.furos_table.setItem(i, 0, QTableWidgetItem(str(furo['diam'])))
            self.furos_table.setItem(i, 1, QTableWidgetItem(str(furo['x'])))
            self.furos_table.setItem(i, 2, QTableWidgetItem(str(furo['y'])))
            delete_btn = QPushButton("Excluir")
            delete_btn.clicked.connect(lambda _, r=i: self.delete_furo_temp(r))
            self.furos_table.setCellWidget(i, 3, delete_btn)
        self.furos_table.resizeColumnsToContents()
    
    def delete_furo_temp(self, row_index):
        if 0 <= row_index < len(self.furos_atuais):
            del self.furos_atuais[row_index]
            self.update_furos_table()

# =============================================================================
# PONTO DE ENTRADA DA APLICAÇÃO
# =============================================================================
def main():
    app = QApplication(sys.argv)
    app.setStyleSheet(INOVA_PROCESS_STYLE)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()